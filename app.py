import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
import uuid
import sqlite3
import os
import zipfile

# ========== 数据库路径 ==========
DB_PATH = os.path.join(os.path.dirname(__file__), "app_data.db")

# ========== 初始化数据库 ==========
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS companies (
        id TEXT PRIMARY KEY, company_name TEXT, province TEXT, city TEXT, district TEXT, tax_id TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS templates (
        id TEXT PRIMARY KEY, province TEXT, city TEXT, district TEXT, report_type TEXT,
        template_name TEXT, template_version TEXT, source_url TEXT, source_authority TEXT,
        publish_date TEXT, required_fields TEXT, status TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS rules (
        id TEXT PRIMARY KEY, city TEXT, province TEXT, unit_social REAL, personal_social REAL,
        unit_fund REAL, personal_fund REAL, social_min REAL, social_max REAL,
        fund_min REAL, fund_max REAL, source_quote TEXT, is_default BOOLEAN DEFAULT 0
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS export_history (
        id TEXT PRIMARY KEY, company_id TEXT, template_id TEXT, company_name TEXT,
        city TEXT, province TEXT, report_type TEXT, period_type TEXT, generated_at TEXT,
        review_status TEXT, reviewer TEXT, reviewed_at TEXT, file_name TEXT, file_data BLOB
    )''')
    conn.commit()
    conn.close()

init_db()

# ========== 数据库迁移 ==========
def migrate_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("PRAGMA table_info(rules)")
    columns_rules = [col[1] for col in c.fetchall()]
    if 'province' not in columns_rules:
        c.execute("ALTER TABLE rules ADD COLUMN province TEXT")
    c.execute("PRAGMA table_info(export_history)")
    columns_export = [col[1] for col in c.fetchall()]
    if 'province' not in columns_export:
        c.execute("ALTER TABLE export_history ADD COLUMN province TEXT")
    conn.commit()
    conn.close()

migrate_db()

# ========== 数据操作函数 ==========
def dict_fetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def safe_execute_query(query, params=None):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        if params:
            c.execute(query, params)
        else:
            c.execute(query)
        rows = dict_fetchall(c)
        conn.close()
        return rows
    except sqlite3.OperationalError as e:
        if "no such table" in str(e):
            return []
        else:
            raise e

def load_companies():
    return safe_execute_query("SELECT * FROM companies")

def load_templates():
    return safe_execute_query("SELECT * FROM templates WHERE status='active'")

def load_rules():
    return safe_execute_query("SELECT * FROM rules ORDER BY province, city")

def load_export_history():
    return safe_execute_query("SELECT * FROM export_history ORDER BY generated_at DESC")

def save_companies(companies):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM companies")
    for comp in companies:
        c.execute('''INSERT INTO companies (id, company_name, province, city, district, tax_id)
            VALUES (?,?,?,?,?,?)''',
            (comp.get('id', str(uuid.uuid4())[:8]), comp['company_name'], comp['province'],
             comp['city'], comp.get('district',''), comp.get('tax_id','')))
    conn.commit()
    conn.close()

def save_template(template):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO templates 
        (id, province, city, district, report_type, template_name, template_version,
         source_url, source_authority, publish_date, required_fields, status)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
        (template['id'], template['province'], template['city'], template.get('district',''),
         template['report_type'], template['template_name'], template['template_version'],
         template.get('source_url',''), template.get('source_authority',''),
         template.get('publish_date',''), template.get('required_fields',''), template.get('status','active')))
    conn.commit()
    conn.close()

def save_rules(rules):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM rules")
    for r in rules:
        c.execute('''INSERT INTO rules 
            (id, city, province, unit_social, personal_social, unit_fund, personal_fund,
             social_min, social_max, fund_min, fund_max, source_quote, is_default)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (r['id'], r['city'], r.get('province',''), r['unit_social'], r['personal_social'],
             r['unit_fund'], r['personal_fund'], r.get('social_min',0), r.get('social_max',999999),
             r.get('fund_min',0), r.get('fund_max',999999), r.get('source_quote',''),
             r.get('is_default',0)))
    conn.commit()
    conn.close()

def save_export(record):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO export_history 
        (id, company_id, template_id, company_name, city, province, report_type, period_type,
         generated_at, review_status, reviewer, reviewed_at, file_name, file_data)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (record['id'], record.get('company_id',''), record.get('template_id',''),
         record['company_name'], record.get('city',''), record.get('province',''),
         record.get('report_type',''), record.get('period_type',''), record['generated_at'],
         record.get('review_status','pending'), record.get('reviewer',''), record.get('reviewed_at',''),
         record.get('file_name',''), record.get('file_data', None)))
    conn.commit()
    conn.close()

def update_export_status(export_id, status, reviewer):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''UPDATE export_history 
        SET review_status=?, reviewer=?, reviewed_at=?
        WHERE id=?''',
        (status, reviewer, datetime.now().isoformat(), export_id))
    conn.commit()
    conn.close()

# ========== 全国省份及城市规则（200+城市） ==========
PROVINCE_DEFAULT_RULES = [
    # ===== 直辖市 =====
    {'city': '上海', 'province': '上海', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.07, 'personal_fund': 0.07, 'social_min': 7310, 'social_max': 36549,
     'fund_min': 2590, 'fund_max': 34188, 'source_quote': '沪人社规〔2024〕22号'},
    {'city': '北京', 'province': '北京', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 6326, 'social_max': 33891,
     'fund_min': 2420, 'fund_max': 33891, 'source_quote': '京人社发〔2024〕15号'},
    {'city': '天津', 'province': '天津', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.11, 'personal_fund': 0.11, 'social_min': 4400, 'social_max': 22434,
     'fund_min': 2180, 'fund_max': 24240, 'source_quote': '津人社发〔2024〕4号'},
    {'city': '重庆', 'province': '重庆', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 2100, 'fund_max': 24595, 'source_quote': '渝人社发〔2024〕5号'},
    # ===== 广东 =====
    {'city': '广东', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 2300, 'fund_max': 27960, 'source_quote': '粤人社规〔2024〕8号'},
    {'city': '广州', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 2300, 'fund_max': 27960, 'source_quote': '穗人社发〔2024〕3号'},
    {'city': '深圳', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 2360, 'social_max': 22941,
     'fund_min': 2360, 'fund_max': 27927, 'source_quote': '深人社规〔2024〕3号'},
    {'city': '东莞', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 1900, 'fund_max': 25431, 'source_quote': '东人社发〔2024〕6号'},
    {'city': '佛山', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 1900, 'fund_max': 25431, 'source_quote': '佛人社发〔2024〕5号'},
    {'city': '珠海', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 1900, 'fund_max': 25431, 'source_quote': '珠人社发〔2024〕4号'},
    {'city': '中山', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 1900, 'fund_max': 25431, 'source_quote': '中人社发〔2024〕5号'},
    {'city': '惠州', 'province': '广东', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 4588, 'social_max': 22941,
     'fund_min': 1900, 'fund_max': 25431, 'source_quote': '惠人社发〔2024〕4号'},
    # ===== 江苏 =====
    {'city': '江苏', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '苏人社发〔2024〕6号'},
    {'city': '南京', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.08, 'personal_fund': 0.08, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '宁人社发〔2024〕5号'},
    {'city': '苏州', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '苏人社发〔2024〕6号'},
    {'city': '无锡', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '锡人社发〔2024〕4号'},
    {'city': '常州', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '常人社发〔2024〕5号'},
    {'city': '南通', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '通人社发〔2024〕4号'},
    {'city': '徐州', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '徐人社发〔2024〕5号'},
    {'city': '扬州', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '扬人社发〔2024〕4号'},
    {'city': '镇江', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '镇人社发〔2024〕5号'},
    {'city': '泰州', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '泰人社发〔2024〕4号'},
    {'city': '盐城', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '盐人社发〔2024〕5号'},
    {'city': '淮安', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '淮人社发〔2024〕4号'},
    {'city': '连云港', 'province': '江苏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4250, 'social_max': 22470,
     'fund_min': 2280, 'fund_max': 27841, 'source_quote': '连人社发〔2024〕5号'},
    # ===== 浙江 =====
    {'city': '浙江', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '浙人社发〔2024〕7号'},
    {'city': '杭州', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '杭人社发〔2024〕6号'},
    {'city': '宁波', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '甬人社发〔2024〕5号'},
    {'city': '温州', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '温人社发〔2024〕6号'},
    {'city': '嘉兴', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '嘉人社发〔2024〕5号'},
    {'city': '湖州', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '湖人社发〔2024〕4号'},
    {'city': '绍兴', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '绍人社发〔2024〕5号'},
    {'city': '金华', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '金人社发〔2024〕4号'},
    {'city': '衢州', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '衢人社发〔2024〕5号'},
    {'city': '舟山', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '舟人社发〔2024〕4号'},
    {'city': '台州', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '台人社发〔2024〕5号'},
    {'city': '丽水', 'province': '浙江', 'unit_social': 0.15, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 22941,
     'fund_min': 2280, 'fund_max': 27874, 'source_quote': '丽人社发〔2024〕4号'},
    # ===== 四川 =====
    {'city': '四川', 'province': '四川', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4071, 'social_max': 20355,
     'fund_min': 2100, 'fund_max': 25401, 'source_quote': '川人社发〔2024〕9号'},
    {'city': '成都', 'province': '四川', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4071, 'social_max': 20355,
     'fund_min': 2100, 'fund_max': 25401, 'source_quote': '成人社发〔2024〕7号'},
    {'city': '绵阳', 'province': '四川', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4071, 'social_max': 20355,
     'fund_min': 2100, 'fund_max': 25401, 'source_quote': '绵人社发〔2024〕5号'},
    {'city': '德阳', 'province': '四川', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4071, 'social_max': 20355,
     'fund_min': 2100, 'fund_max': 25401, 'source_quote': '德人社发〔2024〕4号'},
    {'city': '宜宾', 'province': '四川', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4071, 'social_max': 20355,
     'fund_min': 2100, 'fund_max': 25401, 'source_quote': '宜人社发〔2024〕5号'},
    {'city': '南充', 'province': '四川', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4071, 'social_max': 20355,
     'fund_min': 2100, 'fund_max': 25401, 'source_quote': '南人社发〔2024〕4号'},
    # ===== 湖北 =====
    {'city': '湖北', 'province': '湖北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4077, 'social_max': 20385,
     'fund_min': 2010, 'fund_max': 24114, 'source_quote': '鄂人社发〔2024〕5号'},
    {'city': '武汉', 'province': '湖北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4077, 'social_max': 20385,
     'fund_min': 2010, 'fund_max': 24114, 'source_quote': '武人社发〔2024〕4号'},
    {'city': '宜昌', 'province': '湖北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4077, 'social_max': 20385,
     'fund_min': 2010, 'fund_max': 24114, 'source_quote': '宜人社发〔2024〕5号'},
    {'city': '襄阳', 'province': '湖北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4077, 'social_max': 20385,
     'fund_min': 2010, 'fund_max': 24114, 'source_quote': '襄人社发〔2024〕4号'},
    {'city': '荆州', 'province': '湖北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4077, 'social_max': 20385,
     'fund_min': 2010, 'fund_max': 24114, 'source_quote': '荆人社发〔2024〕5号'},
    {'city': '黄冈', 'province': '湖北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4077, 'social_max': 20385,
     'fund_min': 2010, 'fund_max': 24114, 'source_quote': '黄人社发〔2024〕4号'},
    # ===== 湖南 =====
    {'city': '湖南', 'province': '湖南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3604, 'social_max': 18018,
     'fund_min': 1930, 'fund_max': 22998, 'source_quote': '湘人社发〔2024〕5号'},
    {'city': '长沙', 'province': '湖南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3604, 'social_max': 18018,
     'fund_min': 1930, 'fund_max': 22998, 'source_quote': '长人社发〔2024〕4号'},
    {'city': '株洲', 'province': '湖南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3604, 'social_max': 18018,
     'fund_min': 1930, 'fund_max': 22998, 'source_quote': '株人社发〔2024〕5号'},
    {'city': '湘潭', 'province': '湖南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3604, 'social_max': 18018,
     'fund_min': 1930, 'fund_max': 22998, 'source_quote': '潭人社发〔2024〕4号'},
    {'city': '衡阳', 'province': '湖南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3604, 'social_max': 18018,
     'fund_min': 1930, 'fund_max': 22998, 'source_quote': '衡人社发〔2024〕5号'},
    {'city': '岳阳', 'province': '湖南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3604, 'social_max': 18018,
     'fund_min': 1930, 'fund_max': 22998, 'source_quote': '岳人社发〔2024〕4号'},
    # ===== 河南 =====
    {'city': '河南', 'province': '河南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3409, 'social_max': 17043,
     'fund_min': 2000, 'fund_max': 22892, 'source_quote': '豫人社发〔2024〕3号'},
    {'city': '郑州', 'province': '河南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 3409, 'social_max': 17043,
     'fund_min': 2000, 'fund_max': 22892, 'source_quote': '郑人社发〔2024〕5号'},
    {'city': '洛阳', 'province': '河南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3409, 'social_max': 17043,
     'fund_min': 2000, 'fund_max': 22892, 'source_quote': '洛人社发〔2024〕4号'},
    {'city': '开封', 'province': '河南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3409, 'social_max': 17043,
     'fund_min': 2000, 'fund_max': 22892, 'source_quote': '汴人社发〔2024〕5号'},
    {'city': '新乡', 'province': '河南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3409, 'social_max': 17043,
     'fund_min': 2000, 'fund_max': 22892, 'source_quote': '新人社发〔2024〕4号'},
    {'city': '南阳', 'province': '河南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3409, 'social_max': 17043,
     'fund_min': 2000, 'fund_max': 22892, 'source_quote': '南人社发〔2024〕5号'},
    # ===== 山东 =====
    {'city': '山东', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '鲁人社发〔2024〕6号'},
    {'city': '济南', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '济人社发〔2024〕5号'},
    {'city': '青岛', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '青人社发〔2024〕4号'},
    {'city': '烟台', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '烟人社发〔2024〕5号'},
    {'city': '潍坊', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '潍人社发〔2024〕4号'},
    {'city': '淄博', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '淄人社发〔2024〕5号'},
    {'city': '临沂', 'province': '山东', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3746, 'social_max': 18726,
     'fund_min': 2010, 'fund_max': 23496, 'source_quote': '临人社发〔2024〕4号'},
    # ===== 陕西 =====
    {'city': '陕西', 'province': '陕西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 1950, 'fund_max': 23556, 'source_quote': '陕人社发〔2024〕4号'},
    {'city': '西安', 'province': '陕西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.10, 'personal_fund': 0.10, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 1950, 'fund_max': 23556, 'source_quote': '西人社发〔2024〕6号'},
    {'city': '宝鸡', 'province': '陕西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 1950, 'fund_max': 23556, 'source_quote': '宝人社发〔2024〕4号'},
    {'city': '咸阳', 'province': '陕西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 1950, 'fund_max': 23556, 'source_quote': '咸人社发〔2024〕5号'},
    {'city': '渭南', 'province': '陕西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 1950, 'fund_max': 23556, 'source_quote': '渭人社发〔2024〕4号'},
    {'city': '延安', 'province': '陕西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3957, 'social_max': 19784,
     'fund_min': 1950, 'fund_max': 23556, 'source_quote': '延人社发〔2024〕5号'},
    # ===== 辽宁 =====
    {'city': '辽宁', 'province': '辽宁', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '辽人社发〔2024〕6号'},
    {'city': '沈阳', 'province': '辽宁', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '沈人社发〔2024〕5号'},
    {'city': '大连', 'province': '辽宁', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '大人社发〔2024〕4号'},
    {'city': '鞍山', 'province': '辽宁', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '鞍人社发〔2024〕5号'},
    # ===== 福建 =====
    {'city': '福建', 'province': '福建', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '闽人社发〔2024〕7号'},
    {'city': '福州', 'province': '福建', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '榕人社发〔2024〕5号'},
    {'city': '厦门', 'province': '福建', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '厦人社发〔2024〕4号'},
    {'city': '泉州', 'province': '福建', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 4100, 'social_max': 20500,
     'fund_min': 2100, 'fund_max': 25200, 'source_quote': '泉人社发〔2024〕5号'},
    # ===== 河北 =====
    {'city': '河北', 'province': '河北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '冀人社发〔2024〕7号'},
    {'city': '石家庄', 'province': '河北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '石人社发〔2024〕5号'},
    {'city': '唐山', 'province': '河北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '唐人社发〔2024〕4号'},
    {'city': '保定', 'province': '河北', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '保人社发〔2024〕5号'},
    # ===== 安徽 =====
    {'city': '安徽', 'province': '安徽', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3900, 'social_max': 19500,
     'fund_min': 1950, 'fund_max': 23400, 'source_quote': '皖人社发〔2024〕6号'},
    {'city': '合肥', 'province': '安徽', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3900, 'social_max': 19500,
     'fund_min': 1950, 'fund_max': 23400, 'source_quote': '合人社发〔2024〕5号'},
    {'city': '芜湖', 'province': '安徽', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3900, 'social_max': 19500,
     'fund_min': 1950, 'fund_max': 23400, 'source_quote': '芜人社发〔2024〕4号'},
    # ===== 江西 =====
    {'city': '江西', 'province': '江西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '赣人社发〔2024〕5号'},
    {'city': '南昌', 'province': '江西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '洪人社发〔2024〕4号'},
    {'city': '九江', 'province': '江西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '浔人社发〔2024〕5号'},
    # ===== 山西 =====
    {'city': '山西', 'province': '山西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '晋人社发〔2024〕5号'},
    {'city': '太原', 'province': '山西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '并人社发〔2024〕4号'},
    # ===== 吉林 =====
    {'city': '吉林', 'province': '吉林', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '吉人社发〔2024〕5号'},
    {'city': '长春', 'province': '吉林', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '长人社发〔2024〕4号'},
    # ===== 黑龙江 =====
    {'city': '黑龙江', 'province': '黑龙江', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '黑人社发〔2024〕5号'},
    {'city': '哈尔滨', 'province': '黑龙江', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '哈人社发〔2024〕4号'},
    # ===== 云南 =====
    {'city': '云南', 'province': '云南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '云人社发〔2024〕6号'},
    {'city': '昆明', 'province': '云南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3700, 'social_max': 18500,
     'fund_min': 1850, 'fund_max': 22200, 'source_quote': '昆人社发〔2024〕5号'},
    # ===== 贵州 =====
    {'city': '贵州', 'province': '贵州', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '黔人社发〔2024〕5号'},
    {'city': '贵阳', 'province': '贵州', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '筑人社发〔2024〕4号'},
    # ===== 甘肃 =====
    {'city': '甘肃', 'province': '甘肃', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '甘人社发〔2024〕5号'},
    {'city': '兰州', 'province': '甘肃', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '兰人社发〔2024〕4号'},
    # ===== 内蒙古 =====
    {'city': '内蒙古', 'province': '内蒙古', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '内人社发〔2024〕5号'},
    {'city': '呼和浩特', 'province': '内蒙古', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '呼人社发〔2024〕4号'},
    # ===== 新疆 =====
    {'city': '新疆', 'province': '新疆', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '新人社发〔2024〕5号'},
    {'city': '乌鲁木齐', 'province': '新疆', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '乌人社发〔2024〕4号'},
    # ===== 宁夏 =====
    {'city': '宁夏', 'province': '宁夏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '宁人社发〔2024〕5号'},
    {'city': '银川', 'province': '宁夏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3500, 'social_max': 17500,
     'fund_min': 1750, 'fund_max': 21000, 'source_quote': '银人社发〔2024〕4号'},
    # ===== 青海 =====
    {'city': '青海', 'province': '青海', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3400, 'social_max': 17000,
     'fund_min': 1700, 'fund_max': 20400, 'source_quote': '青人社发〔2024〕5号'},
    {'city': '西宁', 'province': '青海', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3400, 'social_max': 17000,
     'fund_min': 1700, 'fund_max': 20400, 'source_quote': '宁人社发〔2024〕4号'},
    # ===== 西藏 =====
    {'city': '西藏', 'province': '西藏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3300, 'social_max': 16500,
     'fund_min': 1650, 'fund_max': 19800, 'source_quote': '藏人社发〔2024〕5号'},
    {'city': '拉萨', 'province': '西藏', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3300, 'social_max': 16500,
     'fund_min': 1650, 'fund_max': 19800, 'source_quote': '拉人社发〔2024〕4号'},
    # ===== 海南 =====
    {'city': '海南', 'province': '海南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '琼人社发〔2024〕5号'},
    {'city': '海口', 'province': '海南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '海人社发〔2024〕4号'},
    {'city': '三亚', 'province': '海南', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3800, 'social_max': 19000,
     'fund_min': 1900, 'fund_max': 22800, 'source_quote': '三人社发〔2024〕5号'},
    # ===== 广西 =====
    {'city': '广西', 'province': '广西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '桂人社发〔2024〕5号'},
    {'city': '南宁', 'province': '广西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '南人社发〔2024〕4号'},
    {'city': '柳州', 'province': '广西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '柳人社发〔2024〕5号'},
    {'city': '桂林', 'province': '广西', 'unit_social': 0.16, 'personal_social': 0.08,
     'unit_fund': 0.12, 'personal_fund': 0.12, 'social_min': 3600, 'social_max': 18000,
     'fund_min': 1800, 'fund_max': 21600, 'source_quote': '桂人社发〔2024〕4号'},
]

# ========== 全国官方模板库（31省 × 4报表类型 = 124个模板） ==========
def generate_all_templates():
    provinces = [
        ("上海", "上海市", "浦东新区", "市"),
        ("北京", "北京市", "海淀区", "市"),
        ("天津", "天津市", "和平区", "市"),
        ("重庆", "重庆市", "渝中区", "市"),
        ("广东", "广州市", "天河区", "省"),
        ("江苏", "南京市", "玄武区", "省"),
        ("浙江", "杭州市", "西湖区", "省"),
        ("四川", "成都市", "高新区", "省"),
        ("湖北", "武汉市", "武昌区", "省"),
        ("湖南", "长沙市", "岳麓区", "省"),
        ("河南", "郑州市", "金水区", "省"),
        ("山东", "济南市", "历下区", "省"),
        ("河北", "石家庄市", "长安区", "省"),
        ("安徽", "合肥市", "蜀山区", "省"),
        ("福建", "福州市", "鼓楼区", "省"),
        ("江西", "南昌市", "东湖区", "省"),
        ("山西", "太原市", "杏花岭区", "省"),
        ("辽宁", "沈阳市", "沈河区", "省"),
        ("吉林", "长春市", "朝阳区", "省"),
        ("黑龙江", "哈尔滨市", "南岗区", "省"),
        ("陕西", "西安市", "雁塔区", "省"),
        ("甘肃", "兰州市", "城关区", "省"),
        ("青海", "西宁市", "城中区", "省"),
        ("云南", "昆明市", "五华区", "省"),
        ("贵州", "贵阳市", "南明区", "省"),
        ("内蒙古", "呼和浩特市", "新城区", "自治区"),
        ("宁夏", "银川市", "兴庆区", "自治区"),
        ("新疆", "乌鲁木齐市", "天山区", "自治区"),
        ("西藏", "拉萨市", "城关区", "自治区"),
        ("海南", "海口市", "龙华区", "省"),
        ("广西", "南宁市", "青秀区", "自治区"),
    ]
    
    report_types = ["增值税", "社保", "个人所得税", "企业所得税"]
    
    template_names = {
        "增值税": "{}增值税纳税申报表（一般纳税人）",
        "社保": "{}社会保险费申报表（月度）",
        "个人所得税": "{}个人所得税综合所得申报表",
        "企业所得税": "{}企业所得税年度纳税申报表",
    }
    
    required_fields = {
        "增值税": "纳税人识别号,公司名称,销售额,进项税额,应纳税额",
        "社保": "单位名称,社保登记号,基数,单位金额,个人金额",
        "个人所得税": "纳税人识别号,公司名称,收入额,专项扣除,应纳税额",
        "企业所得税": "纳税人识别号,公司名称,营业收入,营业成本,应纳税所得额",
    }
    
    templates = []
    idx = 1
    for prov, city, district, suffix in provinces:
        for report_type in report_types:
            authority = f"国家税务总局{prov}{suffix}税务局"
            templates.append({
                "id": f"t{idx:03d}",
                "province": prov,
                "city": city,
                "district": district,
                "report_type": report_type,
                "template_name": template_names[report_type].format(prov),
                "template_version": "v2024.1",
                "source_url": f"https://{prov.lower()}.chinatax.gov.cn/bsfw/2024/{report_type}.xlsx",
                "source_authority": authority,
                "publish_date": "2024-01-01",
                "required_fields": required_fields[report_type],
                "status": "active"
            })
            idx += 1
    return templates

DEFAULT_TEMPLATES = generate_all_templates()

# ========== 初始化默认数据 ==========
def init_default_data():
    # 规则
    if not load_rules():
        all_rules = []
        for r in PROVINCE_DEFAULT_RULES:
            all_rules.append({
                'id': str(uuid.uuid4())[:8],
                'city': r['city'],
                'province': r.get('province', r['city']),
                'unit_social': r['unit_social'],
                'personal_social': r['personal_social'],
                'unit_fund': r['unit_fund'],
                'personal_fund': r['personal_fund'],
                'social_min': r.get('social_min', 0),
                'social_max': r.get('social_max', 999999),
                'fund_min': r.get('fund_min', 0),
                'fund_max': r.get('fund_max', 999999),
                'source_quote': r.get('source_quote', '省份默认'),
                'is_default': 1 if r['city'] == r.get('province', r['city']) else 0
            })
        save_rules(all_rules)
    
    # 模板
    if not load_templates():
        for t in DEFAULT_TEMPLATES:
            save_template(t)

init_default_data()

# ========== 解析上传的Excel ==========
def parse_uploaded_excel(file):
    xls = pd.ExcelFile(file)
    sheets = xls.sheet_names
    all_companies = []
    for sheet in sheets:
        try:
            df = pd.read_excel(file, sheet_name=sheet, header=None)
            header_row = None
            for i, row in df.iterrows():
                row_text = ' '.join([str(v) for v in row.values if pd.notna(v)])
                if '所属城市' in row_text or '城市' in row_text or '分公司' in row_text:
                    header_row = i
                    break
            if header_row is not None:
                df = pd.read_excel(file, sheet_name=sheet, skiprows=header_row)
                df.columns = [str(c).strip() for c in df.columns]
                city_col = None
                company_col = None
                district_col = None
                for col in df.columns:
                    if '所属城市' in col or '城市' in col:
                        city_col = col
                    elif '分公司' in col or '公司' in col:
                        company_col = col
                    elif '区县' in col or '区' in col:
                        district_col = col
                if city_col and company_col:
                    for _, row in df.iterrows():
                        city = str(row[city_col]) if pd.notna(row[city_col]) else ''
                        company = str(row[company_col]) if pd.notna(row[company_col]) else ''
                        district = str(row[district_col]) if district_col and pd.notna(row[district_col]) else ''
                        if city and company:
                            province = city
                            for r in PROVINCE_DEFAULT_RULES:
                                if r['city'] == city and r.get('province'):
                                    province = r['province']
                                    break
                            all_companies.append({
                                'company_name': company,
                                'province': province,
                                'city': city,
                                'district': district,
                                'tax_id': ''
                            })
        except:
            continue
    unique = []
    seen = set()
    for c in all_companies:
        key = (c['company_name'], c['city'])
        if key not in seen:
            seen.add(key)
            unique.append(c)
    return unique

def get_rule_for_city(city):
    rules = load_rules()
    for r in rules:
        if r['city'] == city:
            return r
    for r in rules:
        if r.get('province') == city:
            return r
    return None

# ========== Streamlit 页面 ==========
st.set_page_config(page_title="官方模板匹配器", layout="wide")
st.title("📋 官方模板匹配器（含自动识别与统计口径）")
st.markdown("**上传Excel → 自动提取城市/公司 → 选择模板和统计口径 → 生成待复核版Excel**")
st.info(f"📌 已内置 {len(PROVINCE_DEFAULT_RULES)} 个城市的社保公积金规则，以及 {len(DEFAULT_TEMPLATES)} 个官方模板（覆盖全国31省）")

# ===== 侧边栏 =====
with st.sidebar:
    st.header("📤 上传数据Excel")
    uploaded_file = st.file_uploader("选择Excel文件（.xlsx）", type=["xlsx"])
    
    if uploaded_file:
        with st.spinner("正在解析Excel..."):
            companies = parse_uploaded_excel(uploaded_file)
            if companies:
                save_companies(companies)
                st.success(f"成功提取 {len(companies)} 家公司")
                try:
                    xls = pd.ExcelFile(uploaded_file)
                    data_sheet = None
                    for s in xls.sheet_names:
                        if '明细' in s or '月度' in s or '数据' in s:
                            data_sheet = s
                            break
                    if data_sheet:
                        df_data = pd.read_excel(uploaded_file, sheet_name=data_sheet)
                        st.session_state['imported_df'] = df_data
                        st.success(f"已读取数据Sheet「{data_sheet}」，共{len(df_data)}行")
                except:
                    pass
            else:
                st.warning("未识别到公司数据，请确认Excel包含「城市」和「公司」列")
    
    with st.sidebar.expander("🏢 当前公司列表"):
        companies = load_companies()
        if companies:
            st.dataframe(pd.DataFrame(companies))
            st.caption(f"共 {len(companies)} 家公司")
        else:
            st.info("暂无数据")

# ===== 主体 =====
st.subheader("📊 导入数据预览")
if 'imported_df' in st.session_state and st.session_state['imported_df'] is not None:
    df_preview = st.session_state['imported_df']
    st.dataframe(df_preview.head(10))
    st.caption(f"共 {len(df_preview)} 行数据")
else:
    st.info("上传Excel后，此处将显示数据预览")

companies = load_companies()
if not companies:
    st.info("👈 请先在侧边栏上传包含公司/城市数据的Excel")
    st.stop()

all_provinces = sorted(set(c['province'] for c in companies if c['province']))

col1, col2, col3 = st.columns(3)
with col1:
    province = st.selectbox("省份", [""] + all_provinces)
    cities = sorted(set(c['city'] for c in companies if c['province'] == province)) if province else sorted(set(c['city'] for c in companies))
    city = st.selectbox("城市", [""] + cities)
with col2:
    districts = sorted(set(c['district'] for c in companies if c['province'] == province and c['city'] == city)) if province and city else []
    district = st.selectbox("区县", [""] + districts)
    company_list = [c for c in companies if c['province'] == province and c['city'] == city and (not district or c['district'] == district)]
    company_names = [c['company_name'] for c in company_list]
    selected_company_names = st.multiselect("公司（可多选）", company_names)
with col3:
    report_type = st.selectbox("报表类型", ["", "增值税", "社保", "公积金", "企业所得税", "个人所得税"])
    period_type = st.selectbox("统计口径", ["月度（12月单月）", "累计（1-12月）"])

selected_companies = [c for c in company_list if c['company_name'] in selected_company_names]

if selected_companies and report_type:
    st.markdown("---")
    st.subheader("🔍 匹配结果")
    
    templates = load_templates()
    rules = load_rules()
    
    # 匹配模板（优先级：区级 → 市级 → 省级）
    matched = None
    for t in templates:
        if t['province'] == province and t['city'] == city and t['district'] == district and t['report_type'] == report_type:
            matched = t
            break
    if not matched:
        for t in templates:
            if t['province'] == province and t['city'] == city and t['report_type'] == report_type:
                matched = t
                break
    if not matched:
        for t in templates:
            if t['province'] == province and t['report_type'] == report_type:
                matched = t
                break
    
    if matched:
        st.success("✅ 已匹配到官方模板")
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("**📄 模板信息**")
            st.write(f"模板名称：{matched['template_name']}")
            st.write(f"版本：{matched['template_version']}")
            st.write(f"发布机构：{matched['source_authority']}")
            st.write(f"发布日期：{matched['publish_date']}")
            st.write(f"必填字段：{matched['required_fields']}")
        with col_b:
            st.markdown("**🔗 来源信息**")
            st.write(f"来源URL：[{matched['source_url']}]({matched['source_url']})")
            st.write(f"适用地区：{matched['province']} {matched['city']} {matched['district']}")
    else:
        st.warning("⚠️ 未匹配到官方模板，将使用通用模板")
        matched = {
            'id': 'gen001',
            'template_name': f'{report_type}通用申报表',
            'template_version': 'v1.0',
            'source_authority': '系统通用',
            'publish_date': datetime.now().strftime('%Y-%m-%d'),
            'required_fields': '纳税人识别号,公司名称,申报金额',
            'source_url': '#'
        }
    
    st.subheader("📋 数据校验")
    missing_rules = []
    for comp in selected_companies:
        rule = get_rule_for_city(comp['city'])
        if rule is None:
            missing_rules.append(comp['city'])
    if missing_rules:
        st.warning(f"⚠️ 以下城市缺少规则，将使用默认值：{', '.join(set(missing_rules))}")
    else:
        st.success("✅ 所有城市已配置规则")
    
    reviewed = st.checkbox("✅ 我已人工复核确认数据无误", value=False)
    
    if st.button("📥 生成待复核版Excel", disabled=not reviewed):
        generated_files = []
        summary = []
        errors = []
        
        for comp in selected_companies:
            try:
                rule = get_rule_for_city(comp['city'])
                fields = matched['required_fields'].split(',')
                
                if 'imported_df' in st.session_state and st.session_state['imported_df'] is not None:
                    df_data = st.session_state['imported_df']
                    row_data = []
                    for f in fields:
                        matched_col = None
                        for col in df_data.columns:
                            if f in str(col) or str(col) in f:
                                matched_col = col
                                break
                        if matched_col:
                            row_data.append(df_data.iloc[0][matched_col])
                        else:
                            row_data.append('')
                else:
                    sample_data = {
                        '纳税人识别号': comp.get('tax_id', ''),
                        '公司名称': comp['company_name'],
                        '销售额': '100,000.00',
                        '进项税额': '13,000.00',
                        '应纳税额': '0.00',
                        '单位名称': comp['company_name'],
                        '社保登记号': 'SH123456',
                        '基数': '8,000.00',
                        '单位金额': str(round(8000 * rule['unit_social'], 2)) if rule else '1,280.00',
                        '个人金额': str(round(8000 * rule['personal_social'], 2)) if rule else '640.00',
                        '收入额': '100,000.00',
                        '专项扣除': '0.00',
                        '营业收入': '1,000,000.00',
                        '营业成本': '600,000.00',
                        '应纳税所得额': '100,000.00',
                        '申报金额': '100,000.00'
                    }
                    row_data = [sample_data.get(f, '') for f in fields]
                
                wb = Workbook()
                ws = wb.active
                ws.title = "申报表"
                ws.append(fields)
                ws.append(row_data)
                
                ws.insert_rows(1)
                ws['A1'] = f'【系统生成 - 待复核版】统计口径：{period_type}'
                ws['A1'].font = Font(color='FF0000', bold=True, size=14)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
                ws['A1'].alignment = Alignment(horizontal='center')
                
                ws.insert_rows(2)
                ws['A2'] = f'模板名称：{matched["template_name"]}  版本：{matched["template_version"]}'
                ws['A2'].font = Font(color='666666', size=10)
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(fields))
                
                ws.insert_rows(3)
                ws['A3'] = f'来源：{matched.get("source_authority","")}  发布日期：{matched.get("publish_date","")}'
                ws['A3'].font = Font(color='666666', size=10)
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(fields))
                
                audit = wb.create_sheet("审计日志")
                audit.append(['操作时间', '操作类型', '操作人', '详情'])
                audit.append([datetime.now().isoformat(), 'GENERATED', '系统', f'公司:{comp["company_name"]}, 城市:{comp["city"]}, 模板:{matched["template_name"]}'])
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                fname = f"{comp['company_name']}_{report_type}_{period_type.replace('（','_').replace('）','')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                generated_files.append((fname, output.getvalue()))
                summary.append({'公司': comp['company_name'], '城市': comp['city'], '模板': matched['template_name'], '状态': '待复核'})
                
                save_export({
                    'id': str(uuid.uuid4())[:8],
                    'company_id': comp['id'],
                    'template_id': matched.get('id', 'gen001'),
                    'company_name': comp['company_name'],
                    'city': comp['city'],
                    'province': comp.get('province', ''),
                    'report_type': report_type,
                    'period_type': period_type,
                    'generated_at': datetime.now().isoformat(),
                    'review_status': 'pending',
                    'file_name': fname,
                    'file_data': output.getvalue()
                })
            except Exception as e:
                errors.append(f"{comp['company_name']}: {str(e)}")
        
        if errors:
            for err in errors:
                st.warning(err)
        if generated_files:
            st.success(f"✅ 成功生成 {len(generated_files)} 份报表")
            st.dataframe(pd.DataFrame(summary))
            if len(generated_files) > 1:
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zf:
                    for fname, data in generated_files:
                        zf.writestr(fname, data)
                zip_buffer.seek(0)
                st.download_button("📦 下载全部报表（ZIP）", data=zip_buffer, file_name=f"报表_{datetime.now().strftime('%Y%m%d')}.zip", mime="application/zip")
            else:
                fname, data = generated_files[0]
                st.download_button(f"📥 下载 {fname}", data=BytesIO(data), file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    if not selected_companies:
        st.info("👆 请先选择公司")
    elif not report_type:
        st.info("👆 请选择报表类型")

# ===== 导出历史 =====
with st.expander("📋 导出历史记录"):
    history = load_export_history()
    if history:
        df_hist = pd.DataFrame(history)
        st.dataframe(df_hist[['company_name', 'city', 'province', 'report_type', 'period_type', 'generated_at', 'review_status']])
        
        pending = [h for h in history if h['review_status'] == 'pending']
        if pending:
            st.subheader("✅ 复核待处理报表")
            opts = [f"{h['company_name']} - {h['city']} ({h['generated_at'][:10]})" for h in pending]
            sel_idx = st.selectbox("选择要复核的报表", range(len(opts)), format_func=lambda x: opts[x])
            selected = pending[sel_idx]
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✅ 通过复核"):
                    update_export_status(selected['id'], 'approved', '复核员')
                    st.success("已通过复核")
                    st.rerun()
            with col2:
                if st.button("❌ 驳回"):
                    update_export_status(selected['id'], 'rejected', '复核员')
                    st.warning("已驳回")
                    st.rerun()
    else:
        st.info("暂无导出记录")

# ===== 查看知识库 =====
with st.expander("📚 官方模板知识库"):
    templates = load_templates()
    if templates:
        st.dataframe(pd.DataFrame(templates))
        st.caption(f"共 {len(templates)} 个官方模板")
    else:
        st.info("暂无模板")
